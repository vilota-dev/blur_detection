import os
import re
import json
import openpyxl

# Define paths
summary_xlsx_path = "/downloads/FLC Failure Summary.xlsx"
output_xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"

# 618D and 619D paths
predictions_paths = {
    "618D": "/dataset_618D/consolidated_batch_predictions.json",
    "619D": "/dataset_619D/consolidated_batch_predictions.json"
}

# 1. Load predictions from both batches
predictions_map = {}  # {(batch, int(SN)): row_data}
all_predictions_raw = []

for batch_name, json_path in predictions_paths.items():
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            predictions = json.load(f)
            print(f"Loaded {len(predictions)} rows from {batch_name} predictions.")
            for row in predictions:
                sn_val = row.get("SN")
                if sn_val is not None:
                    try:
                        predictions_map[(batch_name, int(sn_val))] = row
                        all_predictions_raw.append(row)
                    except ValueError:
                        pass
    else:
        print(f"Warning: predictions path not found: {json_path}")

# 2. Open the original workbook
wb = openpyxl.load_workbook(summary_xlsx_path)
ws = wb["FLC Failure Summary"]

# 3. Write column headers to row 1 (Columns F to P)
headers = [
    (6, "Pos 1 Predict"), (7, "Pos 1 Conf"),
    (8, "Pos 3 Predict"), (9, "Pos 3 Conf"),
    (10, "Pos 5 Predict"), (11, "Pos 5 Conf"),
    (12, "Pos 7 Predict"), (13, "Pos 7 Conf"),
    (14, "Pos 9 Predict"), (15, "Pos 9 Conf"),
    (16, "AI Status"),  # Appending overall AI Status as Column P
]
for col_idx, label in headers:
    ws.cell(row=1, column=col_idx, value=label)

# 4. Populate rows on the main sheet and track matched rows
matched_rows = []
unit_pattern = re.compile(r"^(6\d{2}D)(\d+)$") # Matches 618D01708, 619D01990, etc.

for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val is None:
        continue
    
    val_str = str(val).strip()
    match = unit_pattern.match(val_str)
    if match:
        batch_prefix = match.group(1) # e.g. "618D" or "619D"
        sn_num = int(match.group(2))
        
        # Look up in our loaded predictions map
        key = (batch_prefix, sn_num)
        if key in predictions_map:
            row_data = predictions_map[key]
            matched_rows.append(row_data)
            
            # Write predictions and confidence percentages (columns F to O)
            ws.cell(row=row_idx, column=6, value=row_data.get("pos 1 predict"))
            ws.cell(row=row_idx, column=7, value=row_data.get("pos 1 confidence"))
            ws.cell(row=row_idx, column=8, value=row_data.get("pos 3 predict"))
            ws.cell(row=row_idx, column=9, value=row_data.get("pos 3 confidence"))
            ws.cell(row=row_idx, column=10, value=row_data.get("pos 5 predict"))
            ws.cell(row=row_idx, column=11, value=row_data.get("pos 5 confidence"))
            ws.cell(row=row_idx, column=12, value=row_data.get("pos 7 predict"))
            ws.cell(row=row_idx, column=13, value=row_data.get("pos 7 confidence"))
            ws.cell(row=row_idx, column=14, value=row_data.get("pos 9 predict"))
            ws.cell(row=row_idx, column=15, value=row_data.get("pos 9 confidence"))
            ws.cell(row=row_idx, column=16, value=row_data.get("Status")) # Column P: AI Status

# 5. Insert new tab "AI Detailed Predictions"
detailed_sheet_name = "AI Detailed Predictions"
if detailed_sheet_name in wb.sheetnames:
    ws_detail = wb[detailed_sheet_name]
    wb.remove(ws_detail)
ws_detail = wb.create_sheet(title=detailed_sheet_name)

# Extract headers from predictions structure
if all_predictions_raw:
    detail_headers = list(all_predictions_raw[0].keys())
    ws_detail.append(detail_headers)
    for row_data in matched_rows:
        ws_detail.append([row_data.get(h) for h in detail_headers])

# 6. Save the workbook
wb.save(output_xlsx_path)
print(f"Merge completed! Appended columns and 'AI Status' column. Added {len(matched_rows)} rows to the detailed tab in {output_xlsx_path}")
