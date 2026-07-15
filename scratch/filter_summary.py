import os
import re
import json
import openpyxl

consolidated_json_path = "/dataset/consolidated_batch_predictions.json"
summary_xlsx_path = "/downloads/FLC Failure Summary.xlsx"
output_xlsx_path = "/downloads/FLC Failure Summary_Filtered.xlsx"

# 1. Read consolidated batch predictions to get valid SNs
with open(consolidated_json_path, "r", encoding="utf-8") as f:
    predictions = json.load(f)

# Extract numeric SNs (both as ints and strings)
valid_sns = set()
for row in predictions:
    sn_val = row.get("SN")
    if sn_val is not None:
        try:
            valid_sns.add(int(sn_val))
        except ValueError:
            pass
        valid_sns.add(str(sn_val).strip())

print(f"Loaded {len(valid_sns)} valid SNs from consolidated batch predictions.")

# 2. Open workbook using openpyxl to preserve formatting
wb = openpyxl.load_workbook(summary_xlsx_path)
ws = wb["FLC Failure Summary"]

# 3. Iterate through the sheet from bottom to top to safely delete rows
deleted_count = 0
kept_count = 0
skipped_non_unit_count = 0

# Regular expression to identify unit serial numbers (e.g. 618D01708, 616D00171)
unit_pattern = re.compile(r"^6\d{2}D(\d+)$")

for row_idx in range(ws.max_row, 0, -1):
    val = ws.cell(row=row_idx, column=1).value
    
    if val is None:
        skipped_non_unit_count += 1
        continue
    
    val_str = str(val).strip()
    
    # Check if the string matches a unit format (e.g. 618D01708)
    match = unit_pattern.match(val_str)
    if match:
        sn_num_str = match.group(1)
        sn_int = int(sn_num_str)
        
        # Check if the unit is in consolidated predictions
        if (sn_int in valid_sns) or (sn_num_str in valid_sns):
            kept_count += 1
        else:
            # Not in consolidated, delete the row!
            ws.delete_rows(row_idx)
            deleted_count += 1
    else:
        # Header or summary statistics block, keep it intact
        skipped_non_unit_count += 1

# 4. Save the filtered workbook
wb.save(output_xlsx_path)
print(f"Filtering completed!")
print(f"Kept units: {kept_count}")
print(f"Deleted units: {deleted_count}")
print(f"Skipped non-unit rows (headers/footers/empty): {skipped_non_unit_count}")
print(f"Saved filtered summary to: {output_xlsx_path}")
