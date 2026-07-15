import os
import re
import shutil
import openpyxl

summary_xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"
dest_root = "/downloads/missing_raw_images"
d_drive_620d = "/d_620d"  # Mapped to D:/620D in docker

# 1. Load missing units from Excel
wb = openpyxl.load_workbook(summary_xlsx_path)
ws = wb["FLC Failure Summary"]

missing_620d_sns = []
unit_pattern = re.compile(r"^620D(\d+)$")

for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val is None:
        continue
    
    val_str = str(val).strip()
    match = unit_pattern.match(val_str)
    if match:
        sn_num = int(match.group(1))
        # Column P is AI Status. If it is None or empty, it is missing (not processed yet)!
        ai_status = ws.cell(row=row_idx, column=16).value
        if ai_status is None:
            missing_620d_sns.append(sn_num)

# Unique list
missing_620d_sns = sorted(list(set(missing_620d_sns)))
print(f"Identified {len(missing_620d_sns)} missing 620D units in summary sheet: {missing_620d_sns}")

# 2. Walk through D:/620D exactly ONCE and map filename -> path
print("Indexing directory D:/620D... (this is fast)")
file_map = {}
for root, dirs, files in os.walk(d_drive_620d):
    for f in files:
        file_map[f.lower()] = os.path.join(root, f)
print(f"Indexed {len(file_map)} files.")

# 3. Perform matching and copy files
found_count = 0
not_found_sns = []

for sn in missing_620d_sns:
    sn_found_any = False
    
    # We look for positions 1, 3, 5, 7, 9
    for pos in [1, 3, 5, 7, 9]:
        expected_filename = f"{sn}-{pos}.bmp"
        
        # O(1) Lookup
        file_path_found = file_map.get(expected_filename.lower())
                
        if file_path_found:
            sn_found_any = True
            # Create target folder structure: C:/downloads/missing_raw_images/620D/<POS>/
            target_dir = os.path.join(dest_root, "620D", str(pos))
            os.makedirs(target_dir, exist_ok=True)
            
            dest_file = os.path.join(target_dir, expected_filename)
            shutil.copy2(file_path_found, dest_file)
            found_count += 1
        else:
            print(f"Warning: Could not find raw image for SN {sn} Pos {pos}")
            
    if not sn_found_any:
        not_found_sns.append(sn)

print(f"\nDone! Copied {found_count} raw images to {dest_root}/620D/")
if not_found_sns:
    print(f"Could not find any raw images for SNs: {not_found_sns}")
