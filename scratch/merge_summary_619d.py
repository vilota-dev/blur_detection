import os
import re
import json
import openpyxl

consolidated_json_path = "/dataset/consolidated_batch_predictions.json"
# We read from the previously generated Appended file to keep the 618D data!
summary_xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"
output_xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"

# 1. Load 619D consolidated batch predictions
with open(consolidated_json_path, "r", encoding="utf-8") as f:
    predictions = json.load(f)

# Keep as lookup map {int(SN): row_data}
predictions_map = {}
for row in predictions:
    sn_val = row.get("SN")
    if sn_val is not None:
        try:
            predictions_map[int(sn_val)] = row
        except ValueError:
            pass

# 2. Open the appended workbook
wb = openpyxl.load_workbook(summary_xlsx_path)
ws = wb["FLC Failure Summary"]

# 3. Populate 619D rows on the first sheet
matched_rows = []
unit_pattern = re.compile(r"^619D(\d+)$") # Explicitly match 619D serials!
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val is None:
        continue
    
    val_str = str(val).strip()
    match = unit_pattern.match(val_str)
    if match:
        sn_num = int(match.group(1))
        
        # Look up in consolidated predictions
        if sn_num in predictions_map:
            row_data = predictions_map[sn_num]
            matched_rows.append(row_data)
            
            # Write predictions and confidence percentages (columns F to O)
            ws.cell(row=row_idx, column=6, value=row_data.get("pos 1 predict"))
            ws.cell(row=row_idx, column=7, value=row_data.get("pos 1 confidence"))
            ws.cell(row=row_idx, column=8, value=row_data.get("pos 3 predict"))
            ws.cell(row=row_idx, column=9, value=row_data.get("pos 3 confidence"))
            ws.cell(row=row_idx, column=10, value=row_data.get("pos 5 predict"))
            ws.cell(row=row_idx, column=11, value=row_data.get("pos 5 confidence"))
            ws.cell(row=row_idx, column=12, value=row_data.get("pos 7 predict"))
            ws.cell(row=row_idx, column=13, value=row_data.get("pos 7 confidence"))
            ws.cell(row=row_idx, column=14, value=row_data.get("pos 9 predict"))
            ws.cell(row=row_idx, column=15, value=row_data.get("pos 9 confidence"))

# 4. Append 619D matched rows to the bottom of the "AI Detailed Predictions" tab
detailed_sheet_name = "AI Detailed Predictions"
if detailed_sheet_name not in wb.sheetnames:
    ws_detail = wb.create_sheet(title=detailed_sheet_name)
    # Write headers if it's a new sheet
    if predictions:
        detail_headers = list(predictions[0].keys())
        ws_detail.append(detail_headers)
else:
    ws_detail = wb[detailed_sheet_name]
    # Read existing headers
    detail_headers = [cell.value for cell in ws_detail[1]]

# Append data rows
for row_data in matched_rows:
    ws_detail.append([row_data.get(h) for h in detail_headers])

# 5. Save the workbook
wb.save(output_xlsx_path)
print(f"619D Merge completed! Appended columns and added {len(matched_rows)} rows to the detailed tab in {output_xlsx_path}")
