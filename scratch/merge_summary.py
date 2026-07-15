import os
import re
import json
import openpyxl

consolidated_json_path = "/dataset/consolidated_batch_predictions.json"
summary_xlsx_path = "/downloads/FLC Failure Summary.xlsx"
output_xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"

# 1. Load consolidated batch predictions
with open(consolidated_json_path, "r", encoding="utf-8") as f:
    predictions = json.load(f)

# Keep as lookup map {int(SN): row_data}
predictions_map = {}
for row in predictions:
    sn_val = row.get("SN")
    if sn_val is not None:
        try:
            predictions_map[int(sn_val)] = row
        except ValueError:
            pass

# 2. Open workbook
wb = openpyxl.load_workbook(summary_xlsx_path)
ws = wb["FLC Failure Summary"]

# 3. Write column headers to row 1 (Columns F to O)
headers = [
    (6, "Pos 1 Predict"), (7, "Pos 1 Conf"),
    (8, "Pos 3 Predict"), (9, "Pos 3 Conf"),
    (10, "Pos 5 Predict"), (11, "Pos 5 Conf"),
    (12, "Pos 7 Predict"), (13, "Pos 7 Conf"),
    (14, "Pos 9 Predict"), (15, "Pos 9 Conf"),
]
for col_idx, label in headers:
    ws.cell(row=1, column=col_idx, value=label)

# 4. Populate rows and track matched rows
matched_rows = []
unit_pattern = re.compile(r"^6\d{2}D(\d+)$")
for row_idx in range(2, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val is None:
        continue
    
    val_str = str(val).strip()
    match = unit_pattern.match(val_str)
    if match:
        sn_num = int(match.group(1))
        
        # Look up in consolidated predictions
        if sn_num in predictions_map:
            row_data = predictions_map[sn_num]
            matched_rows.append(row_data)
            
            # Write predictions and confidence percentages
            ws.cell(row=row_idx, column=6, value=row_data.get("pos 1 predict"))
            ws.cell(row=row_idx, column=7, value=row_data.get("pos 1 confidence"))
            ws.cell(row=row_idx, column=8, value=row_data.get("pos 3 predict"))
            ws.cell(row=row_idx, column=9, value=row_data.get("pos 3 confidence"))
            ws.cell(row=row_idx, column=10, value=row_data.get("pos 5 predict"))
            ws.cell(row=row_idx, column=11, value=row_data.get("pos 5 confidence"))
            ws.cell(row=row_idx, column=12, value=row_data.get("pos 7 predict"))
            ws.cell(row=row_idx, column=13, value=row_data.get("pos 7 confidence"))
            ws.cell(row=row_idx, column=14, value=row_data.get("pos 9 predict"))
            ws.cell(row=row_idx, column=15, value=row_data.get("pos 9 confidence"))

# 5. Insert new sheet and write complete detailed rows
detailed_sheet_name = "AI Detailed Predictions"
if detailed_sheet_name in wb.sheetnames:
    ws_detail = wb[detailed_sheet_name]
    wb.remove(ws_detail)
ws_detail = wb.create_sheet(title=detailed_sheet_name)

# Extract headers from predictions structure
if predictions:
    detail_headers = list(predictions[0].keys())
    ws_detail.append(detail_headers)
    for row_data in matched_rows:
        ws_detail.append([row_data.get(h) for h in detail_headers])

# 6. Save the workbook
wb.save(output_xlsx_path)
print(f"Merge completed! Appended columns to FLC Failure Summary and saved detailed tab with {len(matched_rows)} rows.")
