import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import Counter

xlsx_path = "/downloads/FLC Failure Summary_Appended.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
ws_main = wb["FLC Failure Summary"]

# ───────────────────────────────────────────────
# 1. Parse Original Human Summary Statistics
#    (from the SUMMARY STATISTICS section at bottom)
# ───────────────────────────────────────────────
original_stats = {
    "Total SNs Searched":          71,
    "MTF Failures":                 7,
    "FLC Position Failures":       51,
    "Factory Blur / NC Failures":   8,
    "Incomplete / No Data":         4,
    "Pass":                         1,
}

# ───────────────────────────────────────────────
# 2. Parse all data rows from main sheet
# ───────────────────────────────────────────────
data_rows = []
for row_idx in range(2, ws_main.max_row + 1):
    sn = ws_main.cell(row=row_idx, column=1).value
    failure_type = ws_main.cell(row=row_idx, column=3).value
    ai_status = ws_main.cell(row=row_idx, column=16).value

    if ai_status is not None and sn is not None:
        pos_preds = {
            1: ws_main.cell(row=row_idx, column=6).value,
            3: ws_main.cell(row=row_idx, column=8).value,
            5: ws_main.cell(row=row_idx, column=10).value,
            7: ws_main.cell(row=row_idx, column=12).value,
            9: ws_main.cell(row=row_idx, column=14).value,
        }
        data_rows.append({
            "sn": str(sn).strip(),
            "failure_type": str(failure_type).strip() if failure_type else "Unknown",
            "ai_status": str(ai_status).strip(),
            "preds": pos_preds,
        })

ai_total = len(data_rows)
status_counts = Counter(r["ai_status"] for r in data_rows)
ai_pass = status_counts.get("自动通过 (Pass)", 0)
ai_review = status_counts.get("待人工审查 (Review)", 0)

# Map original failure type to AI decisions
fail_type_vs_ai = {}
for r in data_rows:
    ft = r["failure_type"]
    ai = r["ai_status"]
    fail_type_vs_ai.setdefault(ft, Counter())[ai] += 1

# Position stats
pos_stats = {pos: Counter() for pos in [1, 3, 5, 7, 9]}
for r in data_rows:
    for pos, pred in r["preds"].items():
        if pred:
            pos_stats[pos][pred.strip().lower()] += 1

print(f"Loaded {ai_total} AI-evaluated units.")

# ───────────────────────────────────────────────
# Styling Helpers
# ───────────────────────────────────────────────
dark_blue   = "1F4E78"
mid_blue    = "2E75B6"
light_blue  = "BDD7EE"
green_fill  = "E2EFDA"
orange_fill = "FCE4D6"
grey_fill   = "F2F2F2"

def hfill(color):  return PatternFill(start_color=color, end_color=color, fill_type="solid")
def hfont(sz=11, bold=False, color="000000", name="Calibri"):
    return Font(name=name, size=sz, bold=bold, color=color)
def halign(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def side(style="thin", color="D9D9D9"):   return Side(style=style, color=color)
def border(all_sides="thin"):
    s = side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    return Border(left=side("thin"), right=side("thin"), top=side("thin"), bottom=side("medium", "000000"))

def write_header(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hfill(dark_blue); c.font = hfont(11, True, "FFFFFF")
    c.alignment = halign("center"); c.border = border()
    return c

def write_subheader(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = hfill(mid_blue); c.font = hfont(11, True, "FFFFFF")
    c.alignment = halign("center"); c.border = border()
    return c

def write_cell(ws, row, col, value, bold=False, fill=None, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = hfont(11, bold)
    c.alignment = halign(align)
    c.border = border()
    if fill: c.fill = hfill(fill)
    if num_fmt: c.number_format = num_fmt
    return c

# ───────────────────────────────────────────────
# 3. Create Stats Sheet
# ───────────────────────────────────────────────
sname = "AI Summary Statistics"
if sname in wb.sheetnames: del wb[sname]
ws = wb.create_sheet(title=sname)
ws.views.sheetView[0].showGridLines = False

# Title
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 30
tc = ws.cell(row=2, column=2, value="AI Predictions vs Original Statistics Report")
tc.font = hfont(16, True, dark_blue)
tc.alignment = halign("left", "center")

sub = ws.cell(row=3, column=2, value=f"AI Coverage: {ai_total} / {original_stats['Total SNs Searched']} units evaluated  |  Batches: 618D · 619D · 620D · 621D · 622D")
sub.font = hfont(10, False, "595959")
ws.row_dimensions[3].height = 16

# ─────────────────────────────────────────────────────────────────────
# TABLE 1: Original Human Statistics vs AI Coverage Comparison
# ─────────────────────────────────────────────────────────────────────
r = 5
ws.cell(row=r, column=2, value="1.  Original Failure Category vs AI Evaluation Coverage").font = hfont(12, True, dark_blue)
r += 1

for col, txt in [(2, "Failure Category"), (3, "Original Count"), (4, "AI Evaluated"), (5, "Not Yet Evaluated"), (6, "AI Coverage %")]:
    write_header(ws, r, col, txt)
r += 1

# Rows: Original category breakdown (human stats) vs how many AI evaluated
human_categories = [
    ("MTF Failures",                  original_stats["MTF Failures"],                   0),  # No images to evaluate
    ("FLC Position Failures",         original_stats["FLC Position Failures"],           sum(1 for d in data_rows if d["failure_type"] == "FLC Position Fail")),
    ("Factory Blur / NC Failures",    original_stats["Factory Blur / NC Failures"],      sum(1 for d in data_rows if d["failure_type"] in ["Factory Blur", "NC (No Check)"])),
    ("Incomplete / No Data",          original_stats["Incomplete / No Data"],            sum(1 for d in data_rows if d["failure_type"] in ["No Final Result", "NC", "No Data"])),
    ("Pass",                          original_stats["Pass"],                            sum(1 for d in data_rows if d["failure_type"] == "Pass")),
]

for (cat, orig_cnt, ai_cnt) in human_categories:
    not_yet = orig_cnt - ai_cnt
    pct = ai_cnt / orig_cnt if orig_cnt > 0 else 0
    write_cell(ws, r, 2, cat, align="left")
    write_cell(ws, r, 3, orig_cnt)
    write_cell(ws, r, 4, ai_cnt, fill=green_fill if ai_cnt > 0 else None)
    write_cell(ws, r, 5, not_yet, fill=orange_fill if not_yet > 0 else None)
    c = write_cell(ws, r, 6, pct, num_fmt="0%")
    r += 1

# Total row
t1_start_r = 7
write_cell(ws, r, 2, "Total", bold=True, fill=grey_fill)
write_cell(ws, r, 3, original_stats["Total SNs Searched"], bold=True, fill=grey_fill)
write_cell(ws, r, 4, ai_total, bold=True, fill=grey_fill)
write_cell(ws, r, 5, original_stats["Total SNs Searched"] - ai_total, bold=True, fill=grey_fill)
tc2 = write_cell(ws, r, 6, ai_total / original_stats["Total SNs Searched"], bold=True, fill=grey_fill, num_fmt="0%")
for col in range(2, 7):
    ws.cell(row=r, column=col).border = thick_bottom()
r += 3

# ─────────────────────────────────────────────────────────────────────
# TABLE 2: AI Decision Summary vs Comparable Original Stats
# ─────────────────────────────────────────────────────────────────────
ws.cell(row=r, column=2, value="2.  AI Decision Outcomes vs Original Human Decisions").font = hfont(12, True, dark_blue)
r += 1

for col, txt in [(2, "Decision / Category"), (3, "Original (Human)"), (4, "AI Decision"), (5, "Delta"), (6, "Agreement")]:
    write_header(ws, r, col, txt)
r += 1

# Compare: Auto-pass vs original pass, flagged for review vs original failures
comparison_rows = [
    ("Total Evaluated by AI",         ai_total,                   ai_total,   None),
    ("Confirmed Failure (Review)",    original_stats["FLC Position Failures"] + original_stats["Factory Blur / NC Failures"] + original_stats["MTF Failures"],
                                      ai_review,
                                      "Units AI flagged for manual review"),
    ("Clean / Pass",                  original_stats["Pass"],
                                      ai_pass,
                                      f"{ai_pass - original_stats['Pass']:+d} additional AI-identified passes"),
]

for (label, orig_val, ai_val, note) in comparison_rows:
    delta = ai_val - orig_val if orig_val is not None else None
    agree = "✓ Aligned" if delta == 0 else ("↑ More" if delta > 0 else "↓ Fewer") if delta is not None else "-"
    write_cell(ws, r, 2, label, align="left")
    write_cell(ws, r, 3, orig_val)
    write_cell(ws, r, 4, ai_val, fill=green_fill if ai_val >= (orig_val or 0) else orange_fill)
    write_cell(ws, r, 5, delta)
    write_cell(ws, r, 6, agree)
    r += 1

for col in range(2, 7):
    ws.cell(row=r-1, column=col).border = thick_bottom()
r += 2

# ─────────────────────────────────────────────────────────────────────
# TABLE 3: Failure Type vs AI Decision Cross-tabulation
# ─────────────────────────────────────────────────────────────────────
ws.cell(row=r, column=2, value="3.  Original Failure Type vs AI Decision (Cross-tabulation)").font = hfont(12, True, dark_blue)
r += 1

all_ai = sorted(status_counts.keys())
cols_ai = all_ai + ["Total"]
for idx, h in enumerate(["Original Failure Type"] + cols_ai):
    write_header(ws, r, 2 + idx, h)
r += 1

start_t3 = r
for ft, counts in sorted(fail_type_vs_ai.items()):
    write_cell(ws, r, 2, ft, align="left")
    for idx, status in enumerate(all_ai):
        v = counts.get(status, 0)
        fill = green_fill if status.startswith("自动通过") and v > 0 else (orange_fill if v > 0 else None)
        write_cell(ws, r, 3 + idx, v, fill=fill)
    write_cell(ws, r, 3 + len(all_ai), sum(counts.values()), bold=True, fill=grey_fill)
    r += 1

# Total row
write_cell(ws, r, 2, "Total", bold=True, fill=grey_fill)
for idx, status in enumerate(all_ai):
    write_cell(ws, r, 3 + idx, status_counts[status], bold=True, fill=grey_fill)
write_cell(ws, r, 3 + len(all_ai), ai_total, bold=True, fill=grey_fill)
for col in range(2, 4 + len(all_ai)):
    ws.cell(row=r, column=col).border = thick_bottom()
r += 3

# ─────────────────────────────────────────────────────────────────────
# TABLE 4: Position-wise Defect Rates
# ─────────────────────────────────────────────────────────────────────
ws.cell(row=r, column=2, value="4.  Camera Position Defect Distribution (AI Predictions)").font = hfont(12, True, dark_blue)
r += 1

for col, h in enumerate(["Position", "OK (o)", "Blur Failure (f)", "Scratch / Dust (sn)", "Normal (n)", "Defect Rate"]):
    write_header(ws, r, 2 + col, h)
r += 1

for pos in [1, 3, 5, 7, 9]:
    c = pos_stats[pos]
    total_pos = sum(c.values())
    defect = c.get("f", 0) + c.get("sn", 0)
    rate = defect / total_pos if total_pos else 0
    fill = orange_fill if rate > 0.5 else (green_fill if rate == 0 else None)
    write_cell(ws, r, 2, f"Position {pos}")
    write_cell(ws, r, 3, c.get("o", 0))
    write_cell(ws, r, 4, c.get("f", 0), fill=orange_fill if c.get("f", 0) > 0 else None)
    write_cell(ws, r, 5, c.get("sn", 0), fill=orange_fill if c.get("sn", 0) > 30 else None)
    write_cell(ws, r, 6, c.get("n", 0))
    write_cell(ws, r, 7, rate, num_fmt="0.0%", fill=fill, bold=rate > 0.5)
    r += 1

# Total row
all_o  = sum(pos_stats[p].get("o", 0) for p in [1,3,5,7,9])
all_f  = sum(pos_stats[p].get("f", 0) for p in [1,3,5,7,9])
all_sn = sum(pos_stats[p].get("sn",0) for p in [1,3,5,7,9])
all_n  = sum(pos_stats[p].get("n", 0) for p in [1,3,5,7,9])
all_tot = all_o + all_f + all_sn + all_n
write_cell(ws, r, 2, "Total Crops", bold=True, fill=grey_fill)
write_cell(ws, r, 3, all_o,  bold=True, fill=grey_fill)
write_cell(ws, r, 4, all_f,  bold=True, fill=grey_fill)
write_cell(ws, r, 5, all_sn, bold=True, fill=grey_fill)
write_cell(ws, r, 6, all_n,  bold=True, fill=grey_fill)
write_cell(ws, r, 7, (all_f + all_sn) / all_tot if all_tot else 0, bold=True, fill=grey_fill, num_fmt="0.0%")
for col in range(2, 9):
    ws.cell(row=r, column=col).border = thick_bottom()

# ─────────────────────────────────────────────────────────────────────
# Column Widths
# ─────────────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 38
for col_l in ["C", "D", "E", "F", "G", "H"]:
    ws.column_dimensions[col_l].width = 22

# Save
try:
    wb.save(xlsx_path)
    print(f"SUCCESS: Saved to {xlsx_path}")
except PermissionError:
    fallback = xlsx_path.replace(".xlsx", "_Stats.xlsx")
    wb.save(fallback)
    print(f"FALLBACK: File locked, saved to {fallback}")
